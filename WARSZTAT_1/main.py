import re
import pdfplumber
import pandas as pd
from openpyxl import load_workbook

PDFS = [
    "data/FS 4887_BE_06_2025.pdf",
    "data/FS 5555_BE_07_2025.pdf"
]  # oryginalne pliki PDF użytkownika

GODZINY = 56
STAWKA = 70.0
SPRZET = 3200.0
SZABLON = "data/Szablon_Kalkulacja.xlsx"
WYNIK = "output/Kalkulacja_wypelniona.xlsx"

def extract_positions(pdf_paths):
    rows = []
    pat = re.compile(r"^(?P<nazwa>.+?)\s+(?P<ilosc>\d+[\.,]?\d*)\s+(?P<jm>\w+)\s+(?P<cena>\d+[\.,]?\d*)\s+(?P<vat>\d+)\s+(?P<wart>\d+[\.,]?\d*)$")
    for pdf in pdf_paths:
        with pdfplumber.open(pdf) as doc:
            for page in doc.pages:
                txt = page.extract_text() or ""
                for line in txt.splitlines():
                    line = line.replace('\xa0', ' ').replace('.', '')
                    m = pat.search(line)
                    if m:
                        d = m.groupdict()
                        ilosc = float(d['ilosc'].replace(',','.'))
                        cena = float(d['cena'].replace(',','.'))
                        wart = float(d['wart'].replace(',','.'))
                        rows.append([d['nazwa'], ilosc, d['jm'], cena, wart, int(d['vat'])])
    return pd.DataFrame(rows, columns=['nazwa','ilosc','jm','cena_netto','wartosc_netto','vat_proc'])

def write_to_template(df, szablon_path, godziny, stawka, sprzet, out_path):
    materialy_netto = float(df['wartosc_netto'].sum())
    kz = 0.10 * materialy_netto
    robocizna = float(godziny) * float(stawka)

    wb = load_workbook(szablon_path)
    ws_m = wb["Materialy"]
    # dopisz materiały od wiersza 100, żeby nie naruszyć przykładu
    start_row = ws_m.max_row + 1
    for _, r in df.iterrows():
        ws_m.append([r['nazwa'], float(r['ilosc']), r['jm'], float(r['cena_netto']), float(r['wartosc_netto']), int(r['vat_proc'])])

    ws_s = wb["Podsumowanie"]
    # Podmiana wartości
    ws_s["B2"] = materialy_netto
    ws_s["B3"] = kz
    ws_s["B4"] = robocizna
    ws_s["B5"] = sprzet
    ws_s["B6"] = materialy_netto + kz + robocizna + sprzet

    wb.save(out_path)
    return out_path

if __name__ == "__main__":
    df = extract_positions(PDFS)
    print("Zaczytano pozycji:", len(df))
    path = write_to_template(df, SZABLON, GODZINY, STAWKA, SPRZET, WYNIK)
    print("Zapisano:", path)
